#!/usr/bin/env python3
"""
Конвертер CSV в XLSX для Pascal Legacy модуля
"""
import csv
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def convert_csv_to_xlsx(csv_file, xlsx_file):
    """Конвертирует CSV файл в XLSX с форматированием"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Чтение CSV
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Обработка timestamp
                if col_idx == 1 and row_idx > 1:
                    try:
                        dt = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
                        cell.value = dt
                        cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                    except:
                        cell.value = value
                # Обработка логических значений
                elif col_idx == 2 and row_idx > 1:
                    if value == 'ИСТИНА':
                        cell.value = True
                    elif value == 'ЛОЖЬ':
                        cell.value = False
                    else:
                        cell.value = value
                # Обработка чисел
                elif col_idx == 3 and row_idx > 1:
                    try:
                        cell.value = float(value)
                        cell.number_format = '0.00'
                    except:
                        cell.value = value
                # Строки
                else:
                    cell.value = value
                
                # Форматирование заголовка
                if row_idx == 1:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
    
    # Автоматическая ширина столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(xlsx_file)
    print(f"XLSX файл создан: {xlsx_file}")

if __name__ == "__main__":
    # Поиск последнего CSV файла
    output_dir = "output"
    csv_files = [f for f in os.listdir(output_dir) if f.endswith('.csv')]
    
    if csv_files:
        latest_csv = max(csv_files, key=lambda f: os.path.getmtime(os.path.join(output_dir, f)))
        csv_path = os.path.join(output_dir, latest_csv)
        xlsx_path = os.path.join(output_dir, latest_csv.replace('.csv', '.xlsx'))
        convert_csv_to_xlsx(csv_path, xlsx_path)
    else:
        print("CSV файлы не найдены")

